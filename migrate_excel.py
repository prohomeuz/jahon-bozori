"""
Excel → SQLite migration
ABC_yakuniy_narx_jadvali.xlsx faylidagi ma'lumotlarni DB ga import qiladi.
- Mavjud SOLD/RESERVED statuslarni saqlaydi
- Excelda bo'lmagan eski apartamentlarni o'chiradi (bookingsiz)
- Yangi apartamentlarni EMPTY status bilan qo'shadi
- size va notes ni Excel dan yangilaydi
"""
import openpyxl
import sqlite3

EXCEL_PATH = '/Users/mirzoulugbek/Downloads/ABC_yakuniy_narx_jadvali.xlsx'
DB_PATH    = '/Users/mirzoulugbek/Desktop/jahon-bozori/db.sqlite'

# ── 1. Excel ma'lumotlarini o'qish ────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_PATH)

excel_apts = {}   # {id: {block, bolim, floor, size, notes}}

for sheet_name in wb.sheetnames:
    ws    = wb[sheet_name]
    block = sheet_name[0]                          # 'A', 'B', 'C'
    floor = 1 if '1-qavat' in sheet_name else 2

    for row in ws.iter_rows(min_row=4, values_only=True):
        raw_id = row[0]
        if not raw_id or '-' not in str(raw_id):
            continue
        apt_id = str(raw_id).strip()
        parts  = apt_id.split('-')
        if len(parts) < 3:
            continue
        bolim = int(parts[1])

        raw_size  = row[2]
        raw_notes = row[4]

        try:
            size = float(str(raw_size).replace(' ', '').strip()) if raw_size else 0.0
        except ValueError:
            size = 0.0

        notes = str(raw_notes).strip() if raw_notes and str(raw_notes).strip() not in ('None', '') else None

        excel_apts[apt_id] = {
            'id':    apt_id,
            'block': block,
            'bolim': bolim,
            'floor': floor,
            'size':  size,
            'notes': notes,
        }

print(f'Excel dan o\'qildi: {len(excel_apts)} ta apartament')

# ── 2. DB bilan ishlash ───────────────────────────────────────────────────────
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
# Python sqlite3 FK off by default — xavfsiz
cur = conn.cursor()

# Joriy statuslarni saqlash
cur_statuses = {
    row['id']: row['status']
    for row in cur.execute('SELECT id, status FROM apartments').fetchall()
}
print(f'DB da hozir: {len(cur_statuses)} ta apartament')
print(f'  SOLD:     {sum(1 for s in cur_statuses.values() if s == "SOLD")}')
print(f'  RESERVED: {sum(1 for s in cur_statuses.values() if s == "RESERVED")}')

# Bookinglarni muhofaza qilish
booked_ids = {
    row[0]
    for row in cur.execute(
        "SELECT DISTINCT apartment_id FROM bookings WHERE cancelled_at IS NULL"
    ).fetchall()
}
print(f'Active booking ID lari: {booked_ids}')

# ── 3. DB ni yangilash ────────────────────────────────────────────────────────
conn.execute('BEGIN')

inserted = 0
updated  = 0
skipped  = 0

for apt_id, apt in excel_apts.items():
    status = cur_statuses.get(apt_id, 'EMPTY')   # eski status yoki EMPTY

    if apt_id in cur_statuses:
        # Mavjud — size va notes ni yangilaymiz, statusni saqlaymiz
        cur.execute(
            'UPDATE apartments SET size=?, notes=?, block=?, bolim=?, floor=? WHERE id=?',
            (apt['size'], apt['notes'], apt['block'], apt['bolim'], apt['floor'], apt_id)
        )
        updated += 1
    else:
        # Yangi — qo'shamiz
        cur.execute(
            'INSERT INTO apartments (id, block, bolim, floor, size, status, notes) VALUES (?,?,?,?,?,?,?)',
            (apt_id, apt['block'], apt['bolim'], apt['floor'], apt['size'], 'EMPTY', apt['notes'])
        )
        inserted += 1

# DB da bor lekin Excelda yo'q bo'lgan apartamentlarni o'chirish (booking bo'lmaganlarni)
db_only_ids = set(cur_statuses.keys()) - set(excel_apts.keys())
deleted = 0
kept_orphan = 0
for old_id in db_only_ids:
    if old_id in booked_ids:
        print(f'  OGOHLANTIRISH: {old_id} Excelda yo\'q lekin aktiv bookingga ega — saqlanmoqda')
        kept_orphan += 1
    else:
        cur.execute('DELETE FROM apartments WHERE id=?', (old_id,))
        deleted += 1

conn.execute('COMMIT')

# ── 4. Natija ─────────────────────────────────────────────────────────────────
total_now = cur.execute('SELECT COUNT(*) FROM apartments').fetchone()[0]
print()
print('=== Migration natijasi ===')
print(f'  Yangilandi  : {updated}')
print(f'  Qo\'shildi   : {inserted}')
print(f'  O\'chirildi  : {deleted}')
print(f'  Saqlandi*   : {kept_orphan}  (* Excelda yo\'q, lekin booking bor)')
print(f'  Jami hozir  : {total_now}')

# Statuslar tekshiruvi
statuses_now = cur.execute('SELECT status, COUNT(*) FROM apartments GROUP BY status').fetchall()
print(f'  Status taqsimoti: {dict(statuses_now)}')

conn.close()
print('\nMigration muvaffaqiyatli yakunlandi!')
